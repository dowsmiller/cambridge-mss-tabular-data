import os
import pandas as pd
from tqdm import tqdm
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.comments import Comment

# Helper function to save DataFrame as either csv or json file
def save_as(df, output_dir, config_name, format):
    """
    Saves a DataFrame to a file in the specified format.
    Args:
        df (DataFrame): The DataFrame to save.
        output_dir (str): Directory to save the file.
        config_name (str): Name of the configuration file.
        format (str): File format to save as. Must be 'csv' or 'json'.
    """
    os.makedirs(output_dir, exist_ok=True)
    output_filename = f"{config_name}.{format}"
    output_file = os.path.join(output_dir, output_filename)
    try:
        if format == "csv":
            df.to_csv(output_file, index=False, encoding='utf-8-sig')
            tqdm.write(f"Saved '{config_name}' to '{output_file}'")
        elif format == "json":
            df.to_json(output_file, orient='records', lines=True, force_ascii=False)
            tqdm.write(f"Saved '{config_name}' to '{output_file}'")
        else:
            tqdm.write(f"Invalid format '{format}'. Supported formats are 'csv' and 'json'.")
    except Exception as e:
        tqdm.write(f"Saving data to '{output_filename}' failed. Error: {e}")

# Function to save DataFrame list as an xlsx file with individual tables as tabs
def save_as_xlsx(df_list, config_list, output_dir, output_filename):
    """
    Saves a list of DataFrames to an Excel file with each DataFrame in a separate sheet.
    Args:
        df_list (dict): Dictionary of DataFrames to save.
        config_list (dict): Dictionary of configuration DataFrames for headings and sections.
        output_dir (str): Directory to save the Excel file.
        output_filename (str): Name of the output Excel file (without extension).
    """
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f"{output_filename}.xlsx")
    sections_list = [config_list[config_name]['section'].to_numpy() for config_name in config_list.keys()]
    headings_list = [config_list[config_name]['heading'].to_numpy() for config_name in config_list.keys()]
    comments_list = [config_list[config_name]['comment'].to_numpy() for config_name in config_list.keys()]
    tqdm.write(f"Saving '{output_filename}'...")
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Zip over the sheets' data, section titles, and comments
            for (name, df), sections, headings, comments in zip(df_list.items(), sections_list, headings_list, comments_list):
                # Write the DataFrame starting from row 2
                df.to_excel(writer, sheet_name=name, index=False, startrow=1)

                # Set the values in row 2 to headings
                for col_idx, value in enumerate(headings, start=1):
                    writer.sheets[name].cell(row=2, column=col_idx, value=value)

                # Access the workbook and worksheet
                worksheet = writer.sheets[name]

                # Write the section titles into the first row
                for col_idx, value in enumerate(sections, start=1):
                    worksheet.cell(row=1, column=col_idx, value=value)

                # Force text type if value starts with '='
                for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            cell.value = "'" + cell.value

                # Set each column to the width of the content in the second row, with a minimum value
                for col_idx, cell in enumerate(worksheet[2], start=1):
                    column_letter = get_column_letter(col_idx)
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length >= 10:
                            worksheet.column_dimensions[column_letter].width = cell_length
                        else:
                            worksheet.column_dimensions[column_letter].width = 10
                    else:
                        worksheet.column_dimensions[column_letter].width = 10

                # Add comments to each cell of the second row using the relevant value from comments
                for col_idx, comment_text in enumerate(comments, start=1):
                    cell = worksheet.cell(row=2, column=col_idx)
                    comment = Comment(comment_text, "Generated")
                    # Set comment height, assuming 15 characters per line and 15pt per line
                    num_lines = (len(str(comment_text)) // 15)
                    comment.height = 30 + 15 * num_lines
                    cell.comment = comment

                # Set up a filter for each column, with row 2 given as the header value
                last_row = worksheet.max_row
                last_col_letter = get_column_letter(len(sections))
                worksheet.auto_filter.ref = f"A2:{last_col_letter}{last_row}"

                # Merge and center identical consecutive section values in the first row
                merge_and_center_cells(worksheet, sections)

                # Freeze the first two rows
                worksheet.freeze_panes = worksheet['A3']

        tqdm.write(f"Saved data to '{output_filename}'")

    except Exception as e:
        tqdm.write(f"Saving data to '{output_filename}' failed. Error: {e}")

# Helper function to merge and center identical consecutive section values in the first row of an xlsx file
def merge_and_center_cells(worksheet, sections):
    """
    Merges and centers identical consecutive section values in the first row of the worksheet.

    Args:
        worksheet: The worksheet object where merging is applied.
        sections: A list of section titles corresponding to the columns.
    """
    start_col = 1
    for col_idx in range(1, len(sections) + 1):
        if col_idx == len(sections) or sections[col_idx] != sections[start_col - 1]:
            if col_idx - start_col >= 1:
                worksheet.merge_cells(
                    start_row=1, start_column=start_col,
                    end_row=1, end_column=col_idx
                )
                merged_cell = worksheet.cell(row=1, column=start_col)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                worksheet.cell(row=1, column=start_col).alignment = Alignment(horizontal='center', vertical='center')
            start_col = col_idx + 1
