import os
import pandas as pd
import xml.etree.ElementTree as ET
import elementpath
import re
from tqdm import tqdm
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.comments import Comment
from concurrent.futures import ProcessPoolExecutor, as_completed

# Function to read XML files from a directory
def read_xml_files(directory, pattern=".xml"):
    """
    Reads XML files from a specified directory.
    Args:
        directory (str): Directory to search for XML files.
        pattern (str): File extension pattern to match.
    Returns:
        list: List of XML file paths.
    """
    xml_files = []
    try:
        for root, _, files in os.walk(directory):
            for file in files:
                if file.endswith(pattern):
                    xml_files.append(os.path.join(root, file))
        return xml_files
    except Exception as e:
        print(f"Reading XML files in {directory} failed. Error: {e}")
        raise

# Function to parse an XML file and return its root element
def parse_xml(file):
    """
    Parses an XML file and returns the root element.
    Args:
        file (str): Path to the XML file.
    Returns:
        Element: The root element of the parsed XML file.
    """
    try:
        tree = ET.parse(file)
        root = tree.getroot()
        return root
    except Exception as e:
        print(f"Parsing {file} failed. Error: {e}")
        raise

# Function to import all authority files and their config files in parallel.
def import_authority(auth_path="authority", auth_config_path="config/auth"):
    """
    Imports and processes authority files and their configuration in parallel.
    Args:
        auth_path (str): Directory containing authority XML files.
        auth_config_path (str): Directory containing authority configuration CSV files.
    """
    # Read and parse XML authority files in parallel
    authority_files = read_xml_files(auth_path)
    authority = {}
    with ProcessPoolExecutor() as executor:
        futures = {executor.submit(parse_xml, file): file for file in authority_files}
        for future in tqdm(as_completed(futures), total=len(futures), desc="Parsing authority files"):
            file = futures[future]
            try:
                filename = os.path.splitext(os.path.basename(file))[0]
                authority[filename] = future.result()
            except Exception as e:
                print(f"Failed to parse authority file {file}. Error: {e}")

    # Read and parse CSV authority configuration files in parallel
    auth_config_files = sorted(read_xml_files(auth_config_path, pattern=".csv"))
    auth_config_list = {}
    with ProcessPoolExecutor() as executor:
        futures = {executor.submit(pd.read_csv, file, dtype=str): file for file in auth_config_files}
        for future in tqdm(as_completed(futures), total=len(futures), desc="Parsing authority config files"):
            file = futures[future]
            try:
                name = os.path.splitext(os.path.basename(file))[0]
                auth_config_list[name] = future.result()
            except Exception as e:
                print(f"Failed to parse authority config file {file}. Error: {e}")

    # Create an empty DataFrame for each authority configuration file
    auth_df_list = {
        name: pd.DataFrame(columns=config['section'] + ": " + config['heading'])
        for name, config in auth_config_list.items()
    }

    return authority, auth_config_list, auth_df_list

# Function to import all collection files and their config files in parallel.
def import_collection(coll_path="collections", coll_config_path="config/collection"):
    """
    Imports and processes collection files and their configuration in parallel.
    Args:
        coll_path (str): Directory containing collection XML files.
        coll_config_path (str): Directory containing collection configuration CSV files.
    Returns:
        tuple: A dictionary of parsed XML files, a dictionary of configuration DataFrames, and a dictionary of empty DataFrames.
    """
    # Read and parse XML catalogue files in parallel
    catalogue_files = read_xml_files(coll_path)
    catalogue = {}
    with ProcessPoolExecutor() as executor:
        futures = {executor.submit(parse_xml, file): file for file in catalogue_files}
        for future in tqdm(as_completed(futures), total=len(futures), desc="Parsing catalogue files"):
            file = futures[future]
            try:
                filename = os.path.splitext(os.path.basename(file))[0]
                catalogue[filename] = future.result()
            except Exception as e:
                print(f"Failed to parse catalogue file {file}. Error: {e}")

    # Read and parse CSV collection configuration files in parallel
    coll_config_files = sorted(read_xml_files(coll_config_path, pattern=".csv"))
    coll_config_list = {}
    with ProcessPoolExecutor() as executor:
        futures = {executor.submit(pd.read_csv, file, dtype=str): file for file in coll_config_files}
        for future in tqdm(as_completed(futures), total=len(futures), desc="Parsing collection config files"):
            file = futures[future]
            try:
                name = os.path.splitext(os.path.basename(file))[0]
                coll_config_list[name] = future.result()
            except Exception as e:
                print(f"Failed to parse collection config file {file}. Error: {e}")

    # Create an empty DataFrame for each collection configuration file
    coll_df_list = {
        name: pd.DataFrame(columns=config['section'] + ": " + config['heading'])
        for name, config in coll_config_list.items()
    }

    return catalogue, coll_config_list, coll_df_list

# Function to extract data from the authority XML files based on the authority configuration files
def process_authority_file(config_name, config, authority, auth_df_list, bar_pos):
    """
    Processes the authority DataFrame based on the configuration file.
    Args:
        config_name (str): Name of the configuration file.
        config (DataFrame): The configuration DataFrame.
        authority (dict): Dictionary of authority XML files.
        auth_df_list (dict): Dictionary of DataFrames for each configuration file.
    Returns:
        tuple: The configuration name and the processed DataFrame.
    """
    df = auth_df_list[config_name]

    # Extract relevant columns from the configuration file
    auth_files, xpaths = (
        config[col].tolist() for col in ["auth_file", "xpath"]
    )

    # Process each XPath expression in the configuration file
    with ProcessPoolExecutor() as executor:
        futures = {
            executor.submit(process_authority_column, i, xpath, auth_file, authority): i 
            for i, (xpath, auth_file) in enumerate(zip(xpaths, auth_files))
        }
        for future in tqdm(as_completed(futures), total=len(futures), desc=f"File '{config_name}'", position=bar_pos):
            i, results = future.result()
            df.iloc[:, i] = results

    # Defragment the DataFrame
    df = pd.concat([df], ignore_index=True)

    # Optionally sort the DataFrame based on the number after "_" in the first column
    df['temp'] = df.iloc[:, 0].str.extract(r'_(\d+)', expand=False).astype(float)
    df = df.sort_values(by='temp', ascending=True, na_position='last').reset_index(drop=True)
    df.drop(columns='temp', inplace=True)
    
    # Save the DataFrame to a CSV file
    auth_csv_output_dir = "output/auth/csv"
    save_as_csv(df, auth_csv_output_dir, config_name)

    # Save the DataFrame to a JSON file
    auth_json_output_dir = "output/auth/json"
    save_as_json(df, auth_json_output_dir, config_name)

    # Update the DataFrame list with the processed DataFrame
    return config_name, df

# Helper function to process authority columns
def process_authority_column(i, xpath, auth_file, authority):
    """
    Processes a single column for the authority DataFrame.
    Args:
        i (int): The index of the column.
        xpath (str): The XPath expression to extract data.
        auth_file (str): The authority file name.
    Returns:
        tuple: The index and the extracted results.
    """
    auth_xml = authority.get(auth_file)
    results = extract_with_xpath(auth_xml, xpath)
    return i, results

# Function to extract data from the collection XML files based on the collection configuration files
def process_collection_file(config_name, config, catalogue, coll_df_list, auth_df_list, bar_pos):
    """
    Processes the collection DataFrame based on the configuration file.
    Args:
        config_name (str): Name of the configuration file.
        config (DataFrame): The configuration DataFrame.
        catalogue (dict): Dictionary of collection XML files.
        coll_df_list (dict): Dictionary of DataFrames for each configuration file.
        auth_df_list (dict): Dictionary of DataFrames for authority files.
    Returns:
        tuple: The configuration name and the processed DataFrame.
    """
    df = coll_df_list[config_name]

    # Extract relevant columns from the configuration file
    xpaths, auth_files, auth_sections, auth_cols, separators = (
        config[col].tolist() for col in ["xpath", "auth_file", "auth_section", "auth_col", "separator"]
    )

    # Process each XPath expression in the configuration file
    with ProcessPoolExecutor() as executor:
        futures = [
            executor.submit(process_collection_column, i, xpath, auth_file, catalogue, auth_df_list, auth_section, auth_col, separator)
            for i, (xpath, auth_file, auth_section, auth_col, separator) in enumerate(
                zip(xpaths, auth_files, auth_sections, auth_cols, separators)
            )
        ]
        for future in tqdm(as_completed(futures), total=len(futures), desc=f"File '{config_name}'", position=bar_pos):
            i, results = future.result()
            df.iloc[:, i] = results

    # Defragment the DataFrame list
    df = pd.concat([df], ignore_index=True)

    # Optionally sort the DataFrame based on specific columns
    if 'file URL' in df.columns:
        # Extract the numeric part from file URL values
        df['file URL temp'] = df['file URL'].str.extract(r'manuscript_(\d+)')[0].astype(float)
        first_col = df.columns[0]
        # If 'file URL' is the first column, sort only by file URL
        # Otherwise, sort first by file URL then by the first column
        if first_col == 'file URL':
            sort_by = ['file URL temp']
        else:
            sort_by = ['file URL temp', first_col]
        df.sort_values(by=sort_by, ascending=True, na_position='last', inplace=True)
        df.drop(columns=['file URL temp'], inplace=True)
        
    elif 'collection' in df.columns:
        # Use the first column (which is not 'file URL') but extract a number that appears after '_' at the end.
        first_col = df.columns[0]
        df['first_col_num'] = df[first_col].str.extract(r'_(\d+)$')[0].astype(float)
        df.sort_values(by=['collection', 'first_col_num'], ascending=True, na_position='last', inplace=True)
        df.drop(columns=['first_col_num'], inplace=True)
        
    else:
        # If neither 'file URL' nor 'collection' exist, perform a natural sort on the first column.
        first_col = df.columns[0]
        df.sort_values(by=first_col, key=lambda col: col.map(natural_keys), ascending=True, na_position='last', inplace=True)

    # Save the DataFrame to a CSV file
    coll_csv_output_dir = "output/collection/csv"
    save_as_csv(df, coll_csv_output_dir, config_name)

    # Save the DataFrame to a JSON file
    coll_json_output_dir = "output/collection/json"
    save_as_json(df, coll_json_output_dir, config_name)

    # Update the DataFrame list with the processed DataFrame
    return config_name, df

# Helper function to process collection columns
def process_collection_column(i, xpath, auth_file, catalogue, auth_df_list, auth_section, auth_col, separator):
    """
    Processes a collection column by extracting data using XPath and looking up values in an authority file.
    Args:
        i (int): The index of the column.
        xpath (str): The XPath expression to extract data.
        auth_file (str): The authority file name.
        auth_section (str): The section name in the authority file.
        auth_col (str): The column name in the authority file.
        separator (str): The separator for joining values.
    Returns:
        tuple: The index and the processed results.
    """
    # Extract the data using XPath
    results = []
    auth_file = auth_file if pd.notna(auth_file) else None
    auth_df = auth_df_list.get(auth_file) if auth_file else None

    # If no authority file is specified, extract the data and append directly
    if auth_file is None:
        for filename, xml in catalogue.items():
            results.append(extract_with_xpath(xml, xpath))

    # Else extract the data and lookup in the authority DataFrame
    else:
        # Set the separator
        s = get_separator(separator)

        # Set the column name
        col_name = auth_section + ": " + auth_col

        # Lookup the value in the authority file
        for filename, xml in catalogue.items():
            # Build the lookup_data list using the helper function for each data_item
            lookup_data = [
                process_lookup_item(data_item, auth_df, col_name, s)
                for data_item in extract_with_xpath(xml, xpath)
            ]

            results.append(lookup_data)

    # Flatten the results and return
    results = [item for sublist in results for item in (sublist if isinstance(sublist, list) else [sublist])]
    return i, results

# Helper function to apply XPath 2.0 queries to an XML element in the TEI namespace
def extract_with_xpath(xml_element, xpath_expr):
    """
    Extracts data from an XML element using XPath 2.0 queries.
    Args:
        xml_element (Element): The XML element to search.
        xpath_expr (str): The XPath expression to evaluate.
    Returns:
        list: The extracted data.
    """
    try:
        result = elementpath.select(
            xml_element, 
            xpath_expr, 
            namespaces={'tei': 'http://www.tei-c.org/ns/1.0'}
        )
        # Convert non-list results (including booleans) to a list.
        if not isinstance(result, list):
            result = [result]
    except Exception as e:
        print(f"XPath extraction failed. Offending XPath: {xpath_expr}. Error: {e}")
        result = []
    return result

# Helper function to determine the separator for authority lookups
def get_separator(separator):
    separator_map = {
        "space": " ",
        "semi-colon": "; ",
        "comma": ", ",
        "none": ""
    }
    s = separator_map.get(separator)
    if s is None:
        s = "; "
        print(f"Unexpected separator value: {separator}. Using '; ' instead.")
    return s

# Helper function to process data found through the authority lookup
def process_lookup_item(data_item, auth_df, col_name, separator):
    """
    Processes a single data item by looking it up in the authority DataFrame and returning the corresponding value.
    Args:
        data_item (str): The data item to process.
        auth_df (DataFrame): The authority DataFrame.
        col_name (str): The column name in the authority DataFrame.
        separator (str): The separator for joining values.
    Returns:
        str: The processed value, joined by the separator.
    """
    pieces = []
    # Split the data_item on spaces
    for identifier in data_item.split(" "):
        # Filter the DataFrame rows where the first column equals the identifier
        filtered = auth_df[auth_df.iloc[:, 0] == identifier]
        if not filtered.empty:
            # Get the value from the specified column
            value = filtered[col_name].iloc[0]
            # If the value is a boolean, convert its string form to lowercase
            piece = str(value).lower() if isinstance(value, bool) else str(value)
        else:
            piece = ""
        # Add only non-empty strings
        if piece:
            pieces.append(piece)
    
    # Deduplicate preserving the order by leveraging dict.fromkeys
    deduped = list(dict.fromkeys(pieces))
    # If all strings were empty, deduped will be empty; return a single empty string
    return separator.join(deduped) if deduped else ""

# Helper function for natural sorting of strings
def natural_keys(text):
    """
    Convert a string into a list of integers and strings for natural sorting.
    Args:
        text (str): The string to convert.
    Returns:
        list: A list of integers and strings.
    """
    list = [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', text)]
    return list

# Function to save DataFrame as a csv file
def save_as_csv(df, output_dir, config_name):
    """
    Saves a DataFrame to a CSV file.
    Args:
        df (DataFrame): The DataFrame to save.
        output_dir (str): Directory to save the CSV file.
        config_name (str): Name of the configuration file."""
    os.makedirs(output_dir, exist_ok=True)
    output_filename = f"{config_name}.csv"
    output_file = os.path.join(output_dir, output_filename)
    try:
        df.to_csv(output_file, index=False, encoding='utf-8-sig')
        print(f"Saved '{config_name}' to '{output_file}'")
    except Exception as e:
        print(f"Saving data to '{output_filename}' failed. Error: {e}")

# Function to save DataFrame as a json file
def save_as_json(df, output_dir, config_name):
    """
    Saves a DataFrame to a JSON file.
    Args:
        df (DataFrame): The DataFrame to save.
        output_dir (str): Directory to save the JSON file.
        config_name (str): Name of the configuration file.
    """
    os.makedirs(output_dir, exist_ok=True)
    output_filename = f"{config_name}.json"
    output_file = os.path.join(output_dir, output_filename)
    try:
        df.to_json(output_file, orient='records', lines=True, force_ascii=False)
        print(f"Saved '{config_name}' to '{output_file}'")
    except Exception as e:
        print(f"Saving data to '{output_filename}' failed. Error: {e}")

# Function to save DataFrame list as an xlsx file with individual tables as tabs
def save_as_xlsx(df_list, config_list, output_dir, output_filename):
    """
    Saves a list of DataFrames to an Excel file with each DataFrame in a separate sheet.
    Args:
        df_list (dict): Dictionary of DataFrames to save.
        config_list (dict): Dictionary of configuration DataFrames for headings and sections.
        output_dir (str): Directory to save the Excel file.
        output_filename (str): Name of the output Excel file (without extension).
    """
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f"{output_filename}.xlsx")
    sections_list = [config_list[config_name]['section'].to_numpy() for config_name in config_list.keys()]
    headings_list = [config_list[config_name]['heading'].to_numpy() for config_name in config_list.keys()]
    comments_list = [config_list[config_name]['comment'].to_numpy() for config_name in config_list.keys()]
    print(f"Saving '{output_filename}'...")
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Zip over the sheets' data, section titles, and comments
            for (name, df), sections, headings, comments in zip(df_list.items(), sections_list, headings_list, comments_list):
                # Write the DataFrame starting from row 2
                df.to_excel(writer, sheet_name=name, index=False, startrow=1)

                # Set the values in row 2 to headings
                for col_idx, value in enumerate(headings, start=1):
                    writer.sheets[name].cell(row=2, column=col_idx, value=value)

                # Access the workbook and worksheet
                worksheet = writer.sheets[name]

                # Write the section titles into the first row
                for col_idx, value in enumerate(sections, start=1):
                    worksheet.cell(row=1, column=col_idx, value=value)

                # Force text type if value starts with '='
                for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            cell.value = "'" + cell.value

                # Set each column to the width of the content in the second row, with a minimum value
                for col_idx, cell in enumerate(worksheet[2], start=1):
                    column_letter = get_column_letter(col_idx)
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length >= 10:
                            worksheet.column_dimensions[column_letter].width = cell_length
                        else:
                            worksheet.column_dimensions[column_letter].width = 10
                    else:
                        worksheet.column_dimensions[column_letter].width = 10

                # Add comments to each cell of the second row using the relevant value from comments
                for col_idx, comment_text in enumerate(comments, start=1):
                    cell = worksheet.cell(row=2, column=col_idx)
                    comment = Comment(comment_text, "Generated")
                    # Set comment height, assuming 15 characters per line and 15pt per line
                    num_lines = (len(str(comment_text)) // 15)
                    comment.height = 30 + 15 * num_lines
                    cell.comment = comment

                # Set up a filter for each column, with row 2 given as the header value
                last_row = worksheet.max_row
                last_col_letter = get_column_letter(len(sections))
                worksheet.auto_filter.ref = f"A2:{last_col_letter}{last_row}"

                # Merge and center identical consecutive section values in the first row
                merge_and_center_cells(worksheet, sections)

                # Freeze the first two rows
                worksheet.freeze_panes = worksheet['A3']

        print(f"Saved data to '{output_filename}'")

    except Exception as e:
        print(f"Saving data to '{output_filename}' failed. Error: {e}")

# Helper function to merge and center identical consecutive section values in the first row of an xlsx file
def merge_and_center_cells(worksheet, sections):
    """
    Merges and centers identical consecutive section values in the first row of the worksheet.

    Args:
        worksheet: The worksheet object where merging is applied.
        sections: A list of section titles corresponding to the columns.
    """
    start_col = 1
    for col_idx in range(1, len(sections) + 1):
        if col_idx == len(sections) or sections[col_idx] != sections[start_col - 1]:
            if col_idx - start_col >= 1:
                worksheet.merge_cells(
                    start_row=1, start_column=start_col,
                    end_row=1, end_column=col_idx
                )
                merged_cell = worksheet.cell(row=1, column=start_col)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                worksheet.cell(row=1, column=start_col).alignment = Alignment(horizontal='center', vertical='center')
            start_col = col_idx + 1
