import os
import pandas as pd
import xml.etree.ElementTree as ET
import elementpath
from tqdm import tqdm
from itertools import chain
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.comments import Comment

# Function to read XML files from a directory
def read_xml_files(directory, pattern=".xml"):
    xml_files = []
    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith(pattern):
                xml_files.append(os.path.join(root, file))
    return xml_files

# Function to parse an XML file and return its root element
def parse_xml(file):
    tree = ET.parse(file)
    root = tree.getroot()
    return root

# Function to apply XPath 2.0 queries to an XML element in the TEI namespace
def extract_with_xpath(xml_element, xpath_expr):
    result = elementpath.select(xml_element, xpath_expr, namespaces={'tei': 'http://www.tei-c.org/ns/1.0'})
    # Ensure the result is always a list
    if isinstance(result, bool):
        return [result]  # Wrap the boolean in a list
    elif not isinstance(result, list):
        return [result]  # Wrap single values in a list
    return result

# Function to save DataFrame as a csv file
def save_as_csv(df, output_dir, config_name):
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f"{config_name}.csv")
    df.to_csv(output_file, index=False, encoding='utf-8-sig')
    print(f"Saved '{config_name}' to '{output_file}'")

# Function to save DataFrame list as an xlsx file with individual tables as tabs
def save_as_xlsx(df_list, config_list, output_dir, output_filename):
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f"{output_filename}.xlsx")
    sections_list = [config_list[config_name]['section'].to_numpy() for config_name in config_list.keys()]
    comments_list = [config_list[config_name]['comment'].to_numpy() for config_name in config_list.keys()]
    print(f"Saving '{output_filename}'...")
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Zip over the sheets' data, section titles, and comments
            for (name, df), sections, comments in zip(df_list.items(), sections_list, comments_list):
                # Write the DataFrame starting from row 2 (leaves row 1 for the section titles)
                df.to_excel(writer, sheet_name=name, index=False, startrow=1)

                # Access the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets[name]

                # Write the section titles into the first row
                for col_idx, value in enumerate(sections, start=1):
                    worksheet.cell(row=1, column=col_idx, value=value)

                # Force text type if value starts with '='
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            cell.value = "'" + cell.value

                # Set each column to the width of the content in the second row, plus some padding
                for col_idx, cell in enumerate(worksheet[2], start=1):
                    column_letter = get_column_letter(col_idx)
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        worksheet.column_dimensions[column_letter].width = cell_length + 2
                    else:
                        worksheet.column_dimensions[column_letter].width = 10

                # Add comments to each cell of the second row using the relevant value from comments
                for col_idx, comment_text in enumerate(comments, start=1):
                    cell = worksheet.cell(row=2, column=col_idx)
                    comment = Comment(comment_text, "Generated")
                    # Set comment height, assuming 15 characters per line and 15pt per line
                    num_lines = (len(str(comment_text)) // 15)
                    comment.height = 30 + 15 * num_lines
                    cell.comment = comment

                # Set up a filter for each column, with row 2 given as the header value
                last_row = worksheet.max_row
                last_col_letter = get_column_letter(len(sections))
                worksheet.auto_filter.ref = f"A2:{last_col_letter}{last_row}"

                # Merge and centre identical consecutive section values in the first row
                start_col = 1
                for col_idx in range(1, len(sections) + 1):
                    if col_idx == len(sections) or sections[col_idx] != sections[start_col - 1]:
                        if col_idx - start_col >= 1:
                            worksheet.merge_cells(
                                start_row=1, start_column=start_col,
                                end_row=1, end_column=col_idx
                            )
                            merged_cell = worksheet.cell(row=1, column=start_col)
                            merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            worksheet.cell(row=1, column=start_col).alignment = Alignment(horizontal='center', vertical='center')
                        start_col = col_idx + 1

                # Freeze the first two rows
                worksheet.freeze_panes = worksheet['A3']

        print(f"Saved collection data to '{output_filename}'")

    except Exception as e:
        print(f"Saving collection data to '{output_filename}' failed.")

# Step 1: Read and parse XML authority files
authority_files = read_xml_files("authority")
authority = {}
for file in tqdm(authority_files, desc="Parsing authority files"):
    filename = os.path.splitext(os.path.basename(file))[0]
    authority[filename] = parse_xml(file)

# Step 2: Read and parse XML catalogue files
catalogue_files = read_xml_files("collections")
catalogue = {}
for file in tqdm(catalogue_files, desc="Parsing catalogue files"):
    filename = os.path.splitext(os.path.basename(file))[0]
    catalogue[filename] = parse_xml(file)

# Step 3: Read and parse CSV authority configuration files
auth_config_files = sorted(read_xml_files("config/auth", pattern=".csv"))
auth_config_list = {}
for file in tqdm(auth_config_files, desc="Parsing authority config files"):
    name = os.path.splitext(os.path.basename(file))[0]
    auth_config_list[name] = pd.read_csv(file, dtype=str)

# Step 4: Read and parse CSV collection configuration files
coll_config_files = sorted(read_xml_files("config/collection", pattern=".csv"))
coll_config_list = {}
for file in tqdm(coll_config_files, desc="Parsing collection config files"):
    name = os.path.splitext(os.path.basename(file))[0]
    coll_config_list[name] = pd.read_csv(file, dtype=str)

# Step 5: Create an empty DataFrame for each authority configuration file
auth_df_list = {
    name: pd.DataFrame(columns=config['heading'].tolist())
    for name, config in auth_config_list.items()
}

# Step 6: Create an empty DataFrame for each collection configuration file
coll_df_list = {
    name: pd.DataFrame(columns=config['heading'].tolist())
    for name, config in coll_config_list.items()
}

# Step 7: Extract data from the authority XML files based on the authority configuration files
for config_name, config in tqdm(auth_config_list.items(), desc="Authority progress"):
    df = auth_df_list[config_name]

    # Step 7.1 Extract relevant columns from the configuration file
    headings, auth_files, xpaths = (
        config[col].tolist() for col in ["heading", "auth_file", "xpath"]
    )

    # Step 7.2: Process each XPath expression in the configuration file
    for xpath, heading, auth_file in tqdm(zip(xpaths, headings, auth_files), total=len(xpaths), desc=f"File '{config_name}'"):
        auth_xml = authority.get(auth_file)
        df[heading] = extract_with_xpath(auth_xml, xpath)

    # Step 7.3: Optionally sort the DataFrame based on the number after "_" in the first column
    df['temp'] = df.iloc[:, 0].str.extract(r'_(\d+)', expand=False).astype(float)
    df = df.sort_values(by='temp', ascending=True, na_position='last').reset_index(drop=True)
    df.drop(columns='temp', inplace=True)

    # Step 7.4: Defragment the DataFrame
    df = pd.concat([df], ignore_index=True)
    
    # Step 7.5: Save the DataFrame to a CSV file
    auth_csv_output_dir = "output/auth/csv"
    save_as_csv(df, auth_csv_output_dir, config_name)

    # Step 7.6: Update the DataFrame list with the processed DataFrame
    auth_df_list[config_name] = df

# Step 8: Save the DataFrame list to an .xlsx file with separate tabs
auth_xlsx_output_dir = "output/auth"
auth_output_filename = "authority_data"
save_as_xlsx(auth_df_list, auth_config_list, auth_xlsx_output_dir, auth_output_filename)

# Step 9: Extract data from the collection XML files based on the collection configuration files
for config_name, config in tqdm(coll_config_list.items(), desc="Collections progress"):
    df = coll_df_list[config_name]

    # Step 9.1: Extract relevant columns from the configuration file
    headings, xpaths, auth_files, auth_cols = (
        config[col].tolist() for col in ["heading", "xpath", "auth_file", "auth_col"]
    )

    # Step 9.2: Process each XPath expression in the configuration file
    for xpath, heading, auth_file, auth_col in tqdm(zip(xpaths, headings, auth_files, auth_cols), total=len(xpaths), desc=f"File '{config_name}'"):
        results = []
        auth_file = auth_file if pd.notna(auth_file) else None
        auth_df = auth_df_list.get(auth_file) if auth_file else None

        # If no authority file is specified, extract the data and append directly
        if auth_file is None:
            for filename, xml in tqdm(catalogue.items(), total=len(catalogue), desc=f"Column '{heading}'", leave=False):
                results.append(extract_with_xpath(xml, xpath))

        # Else extract the data and lookup in the authority DataFrame
        else:
            for filename, xml in tqdm(catalogue.items(), total=len(catalogue), desc=f"Column '{heading}' (authority lookup)", leave=False):
                lookup_data = [
                    "; ".join(
                        str(auth_df[auth_df.iloc[:, 0] == identifier][auth_col].iloc[0])
                        if not auth_df[auth_df.iloc[:, 0] == identifier].empty else ""
                        for identifier in data_item.split(" ")
                    )
                    for data_item in extract_with_xpath(xml, xpath)
                ]

                results.append(lookup_data)

        # Flatten the results and add them to the DataFrame
        results = [item for sublist in results if isinstance(sublist, list) for item in sublist]
        df[heading] = results

    # Step 9.3: Optionally sort the DataFrame based on specific columns
    if 'file URL' in df.columns:
        df['file URL temp'] = df['file URL'].str.extract(r'manuscript_(\d+)')[0].astype(float)
        df.sort_values(by=['file URL temp'], ascending=True, na_position='last', inplace=True)
        df.drop(columns=['file URL temp'], inplace=True)
    elif 'item UID' in df.columns:
        df['item UID temp'] = df['item UID']
        df.sort_values(by=['item UID temp'], ascending=True, na_position='last', inplace=True)
        df.drop(columns=['item UID temp'], inplace=True)
    else:
        print(f"Warning: no sorting column found in '{config_name}'. Skipping sorting.")

    # Step 9.4: Defragment the DataFrame list
    df = pd.concat([df], ignore_index=True)

    # Step 9.5: Save the DataFrame to a CSV file
    coll_csv_output_dir = "output/collection/csv"
    save_as_csv(df, coll_csv_output_dir, config_name)

    # Step 9.6: Update the DataFrame list with the processed DataFrame
    coll_df_list[config_name] = df

# Step 10: Save the DataFrame list to an .xlsx file with separate tabs
coll_xlsx_output_dir = "output/collection"
coll_output_filename = "collection_data"
save_as_xlsx(coll_df_list, coll_config_list, coll_xlsx_output_dir, coll_output_filename)